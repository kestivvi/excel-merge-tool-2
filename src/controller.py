from copy import Error, error
import datetime
from os import path
from model import Model
from view_abc import View, BtnStatus

import openpyxl as xl 

class Controller:

    def __init__(self, model: Model, view: View):
        self.model = model
        self.view = view


    def handle_choose_input_directory_click(self):
        inputDirectory = self.view.askForInputDirectory()

        errors = []
        for file in self.model.files_to_check:
            full_path = inputDirectory + '/' + file
            if not path.exists(full_path):
                errors.append(f"File {full_path} does not exists!")
        
        if not errors == []:
            self.view.alert_error("\n".join(errors))
            return

        self.model.inputDirectory = inputDirectory
        self.view.setInputDirectoryText(self.model.inputDirectory)

        self.model.path_to_save = self.model.inputDirectory + "/Directors Basic working file - output.xlsm"
        self.view.setSaveLocationText(self.model.path_to_save)

        self.view.checkBtnMergeStatus()
        
    def get_column_mapping(self, ws, row_with_columns=1):
        columns = {}
        index = 1
        column_title = ws.cell(row=row_with_columns, column=index).value.strip()
        while True:
            columns[column_title] = index
            index += 1
            cell_value = ws.cell(row=row_with_columns, column=index).value
            if cell_value is None or cell_value == "":
                break
            column_title = ws.cell(row=row_with_columns, column=index).value.strip()
        return columns

    def copy_simple_case(self, ws_out, ws_in, ws_in_starting_row=2):
        output_columns = self.get_column_mapping(ws_out)
        input_columns = self.get_column_mapping(ws_in)

        row_in = ws_in_starting_row
        row_out = 2
        row_in_max = ws_in.max_row
        while row_in <= row_in_max:
            if ws_in.cell(row=row_in, column=input_columns["Title Description"]).value.strip() == "65 - 65 (BY CONVERSION)":
                row_in += 1
                continue
                
            for column, index in output_columns.items():
                input_value = ws_in.cell(row=row_in, column=input_columns[column]).value
                ws_out.cell(row=row_out, column=index).value = input_value
            
            row_in += 1 
            row_out += 1 


    def copy_recruits(self, ws_in, ws_recruits, ws_nofos):
        columns_in = self.get_column_mapping(ws_in)
        columns_recruits = self.get_column_mapping(ws_recruits)
        columns_nofos = self.get_column_mapping(ws_nofos)

        row_in = 2
        row_recruits = 2
        row_nofos = 2
        row_in_max = ws_in.max_row
        while row_in <= row_in_max:
            if ws_in.cell(row=row_in, column=columns_in["Title Description"]).value.strip() == "65 - 65 (BY CONVERSION)":
                row_in += 1
                continue
            
            temp_columns = columns_recruits
            temp_ws = ws_recruits
            temp_row = row_recruits
            nofos = ws_in.cell(row=row_in, column=columns_in["BP per Product"]).value == 1
            if nofos:
                temp_columns = columns_nofos
                temp_ws = ws_nofos
                temp_row = row_nofos
                
            for column, index in temp_columns.items():
                input_value = ws_in.cell(row=row_in, column=columns_in[column]).value
                temp_ws.cell(row=temp_row, column=index).value = input_value

            row_in += 1

            if nofos:
                row_nofos += 1
            else:
                row_recruits += 1
    
    def Make_TitlesReports_Mapping(self, TitlesReportMatureMarketsWB):
        mapping = {}
        # 98123: {
        #   "Member group": "Guest Customer",
        #   "Director First Name": "JACKIE",
        #   "Director Last Name": "VENABLES",
        # }
        ws = TitlesReportMatureMarketsWB["C1-M2 titles all"]
        columns = self.get_column_mapping(ws, row_with_columns=11)
        
        for row in range(12, ws.max_row):
            consultant_number = ws.cell(row=row, column=columns["Consultant (Downline)"]).value
            if consultant_number not in mapping:
                member_group = ws.cell(row=row, column=columns["Legacy"]).value
                mapping[consultant_number] = {}
                mapping[consultant_number]["Member group"] = member_group

                director_first_name = ws.cell(row=row, column=columns["Director First Name"]).value
                mapping[consultant_number]["Director First Name"] = director_first_name
                director_last_name = ws.cell(row=row, column=columns["Director Last Name"]).value
                mapping[consultant_number]["Director Last Name"] = director_last_name


        ws = TitlesReportMatureMarketsWB["VIP with BP >0"]
        columns = self.get_column_mapping(ws, row_with_columns=12)

        for row in range(13, ws.max_row):
            consultant_number = ws.cell(row=row, column=columns["End Customer"]).value
            if consultant_number not in mapping:
                member_group = ws.cell(row=row, column=columns["End Customer Type"]).value
                mapping[consultant_number] = {}
                mapping[consultant_number]["Member group"] = member_group

                director_first_name = ws.cell(row=row, column=columns["Director First Name"]).value
                mapping[consultant_number]["Director First Name"] = director_first_name
                director_last_name = ws.cell(row=row, column=columns["Director Last Name"]).value
                mapping[consultant_number]["Director Last Name"] = director_last_name
        
        return mapping

    def copy_skincare_case(self, ws_out, ws_in, mapping):
        output_columns = self.get_column_mapping(ws_out)
        input_columns = self.get_column_mapping(ws_in)

        row_in = 2
        row_out = 2
        row_in_max = ws_in.max_row
        while row_in <= row_in_max:
            if ws_in.cell(row=row_in, column=input_columns["Title Description"]).value.strip() == "65 - 65 (BY CONVERSION)":
                row_in += 1
                continue
                
            for column, index in output_columns.items():
                if column not in input_columns:
                    continue

                input_value = ws_in.cell(row=row_in, column=input_columns[column]).value
                ws_out.cell(row=row_out, column=index).value = input_value

                if column == "Consultant number":
                    if input_value in mapping:
                        member_group = mapping[input_value]["Member group"]
                        ws_out.cell(row=row_out, column=output_columns["Member group"]).value = member_group
                    
            row_in += 1 
            row_out += 1


    def copy_catalogue_titles_case(self, ws_out1, ws_out2, wb_in):
        ws_in = wb_in["C1-M2 titles all"]

        output_columns = self.get_column_mapping(ws_out1)
        input_columns = self.get_column_mapping(ws_in, row_with_columns=11)

        row_out = 2
        for row in range(12, ws_in.max_row+1):
                
            for column, index in output_columns.items():
                input_value = ws_in.cell(row=row, column=input_columns[column]).value
                ws_out1.cell(row=row_out, column=index).value = input_value
                ws_out2.cell(row=row_out, column=index).value = input_value
            
            row_out += 1
            
    
    def copy_YTD_case(self, ws_out, ws_in, mapping): 
        output_columns = self.get_column_mapping(ws_out)
        input_columns = self.get_column_mapping(ws_in)

        for row in range(2, ws_in.max_row+1):    
            for column, index in output_columns.items():
                
                if column not in input_columns:
                    continue

                input_value = ws_in.cell(row=row, column=input_columns[column]).value
                ws_out.cell(row=row, column=index).value = input_value

                if column == "CONSULTANT":
                    if input_value in mapping:
                        director_first_name = mapping[input_value]["Director First Name"]
                        ws_out.cell(row=row, column=output_columns["DIRECTOR_FIRST_NAME"]).value = director_first_name
                        director_last_name = mapping[input_value]["Director Last Name"]
                        ws_out.cell(row=row, column=output_columns["DIRECTOR_LAST_NAME"]).value = director_last_name
    

    def copy_VIP_Recruits_case(self, ws_out, ws_in, date_from, date_to):

        output_columns = self.get_column_mapping(ws_out)
        input_columns = self.get_column_mapping(ws_in, row_with_columns=12)

        row_out = 2
        row_in = 13
        row_in_max = ws_in.max_row
        while row_in <= row_in_max:
            
            end_customer_type = ws_in.cell(row=row_in, column=input_columns["End Customer Type"]).value
            if end_customer_type != "VIP Customer":
                row_in += 1
                continue
            
            order_date = ws_in.cell(row=row_in, column=input_columns["End customers first order date"]).value
            if not (date_from <= order_date and order_date <= date_to):
                row_in += 1
                continue

            for column, index in output_columns.items():
                input_value = ws_in.cell(row=row_in, column=input_columns[column]).value
                ws_out.cell(row=row_out, column=index).value = input_value
            
            row_in += 1
            row_out += 1


    def checkDates(self):
        def check(year, moth, day):
            if month < 1:
                raise ValueError("Month cannot be lower than 1")
            if month > 12:
                raise ValueError("Month cannot be greater than 12")
            if day < 1:
                raise ValueError("Day cannot be lower than 1")
            if day > 31:
                raise ValueError("Day cannot be greater than 31")
            
        try:
            date_from = str(self.view.getStartDate()).strip().split("-") # datetime.datetime(2021, 8, 15)
            year = int(date_from[0])
            month = int(date_from[1])
            day = int(date_from[2])
            
            check(year, month, day)
            date_from = datetime.datetime(year, month, day)


            date_to = str(self.view.getEndDate()).strip().split("-") # datetime.datetime(2021, 9, 7)
            year = int(date_to[0])
            month = int(date_to[1])
            day = int(date_to[2])
            
            check(year, month, day)
            date_to = datetime.datetime(int(date_to[0]), int(date_to[1]), int(date_to[2]))

            if date_from > date_to:
                raise ValueError("The Start Date cannot be greater than the End Date!")

            return True
        except Exception as error:
            self.view.alert_error(f"Dates should be in format YYYY-MM-DD!\nERROR: {error}")
            return False



    def handle_merge_click(self):

        if not self.checkDates():
            return

        self.view.setBtnMergeStatus(BtnStatus.DISABLED)

        self.view.setProgressText("Loading Directors workbook...")
        
        output_workbook = xl.load_workbook(self.model.inputDirectory + "/Directors Basic working file.xlsm", keep_vba=True)

        ######################################
        ### Copy simple cases

        simple_cases = [
            "Signups",
            "Starter Kits",
            "Welcome Programme",
            "I3"
        ]

        for case in simple_cases:
            self.view.setProgressText(case)
            wb_in = xl.load_workbook(self.model.inputDirectory + f"/{case}.xlsx")
            self.copy_simple_case(output_workbook[case], wb_in.active)
        

        ######################################
        ### Copy "Recruits (FO)" and "NOFOs who paid joining fee" case

        self.view.setProgressText("Recruits (FO) and NOFOs")
        wb_in = xl.load_workbook(self.model.inputDirectory + "/Recruits (FO).xlsx")
        self.copy_recruits(wb_in.active, output_workbook["Recruits (FO)"], output_workbook["NOFOs who paid joining fee"])
        
        ######################################
        ### Copy "Skincare Sets" case
        
        self.view.setProgressText("Skincare Sets")
        # Make "Consultant number - Member group" mapping
        titles_reports_wb = xl.load_workbook(self.model.inputDirectory + "/Titles Report Mature Markets.xlsx")
        mapping = self.Make_TitlesReports_Mapping(titles_reports_wb)

        wb_in = xl.load_workbook(self.model.inputDirectory + "/Skincare Sets.xlsx")
        self.copy_skincare_case(output_workbook["Skincare Sets"], wb_in.active, mapping)

        
        ######################################
        ### Copy "Catalogue BP Sales" oraz "Titles" cases

        self.view.setProgressText("Catalogue BP Sales and Titles")
        self.copy_catalogue_titles_case(
            output_workbook["Catalogue BP Sales"],
            output_workbook["Titles"],
            titles_reports_wb
        )


        ######################################
        ### Copy "YTD"

        self.view.setProgressText("YTD")
        ws_in = xl.load_workbook(self.model.inputDirectory + "/YTD.xlsx")
        self.copy_YTD_case(output_workbook["YTD"], ws_in.active, mapping)


        ######################################
        ### Copy "VIP Recruits"

        self.view.setProgressText("VIP Recruits")

        date_from = str(self.view.getStartDate()).strip().split("-") # datetime.datetime(2021, 8, 15)
        date_from = datetime.datetime(int(date_from[0]), int(date_from[1]), int(date_from[2]))

        date_to = str(self.view.getEndDate()).strip().split("-") # datetime.datetime(2021, 9, 7)
        date_to = datetime.datetime(int(date_to[0]), int(date_to[1]), int(date_to[2]))

        self.copy_VIP_Recruits_case(output_workbook["VIP Recruits"], titles_reports_wb["VIP with BP >0"], date_from, date_to)



        self.view.setProgressText("Saving now...")
        output_workbook.save(self.model.path_to_save)
        self.view.setProgressText("Done!")
        self.view.notifySound()

        self.view.checkBtnMergeStatus()
    

    def start(self):
        self.view.setUpView(self)
        self.view.startMainLoop()
