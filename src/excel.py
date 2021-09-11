from datetime import datetime
import openpyxl as xl

from typing import Callable, List


def get_column_mapping(ws, row_with_columns=1, starting_column=1) -> dict:
    columns = {}
    col_num = starting_column
    column_title = ws.cell(row=row_with_columns, column=col_num).value
    
    while column_title is not None and column_title != "":
        column_title = column_title.strip()
        columns[column_title] = col_num
        col_num += 1
        column_title = ws.cell(row=row_with_columns, column=col_num).value
    
    return columns


def copy_simple_case(ws_out, ws_in, ws_in_starting_row=2) -> None:
    col_out = get_column_mapping(ws_out)
    col_in = get_column_mapping(ws_in)
    
    row_in = ws_in_starting_row
    row_out = 2
    row_in_max = ws_in.max_row
    
    while row_in <= row_in_max:
        title_description = ws_in.cell(row=row_in, column=col_in["Title Description"]).value.strip()
        if title_description == "65 - 65 (BY CONVERSION)":
            row_in += 1
            continue
            
        for column, index in col_out.items():
            input_value = ws_in.cell(row=row_in, column=col_in[column]).value
            ws_out.cell(row=row_out, column=index).value = input_value
        
        row_in += 1 
        row_out += 1 


def copy_recruits(ws_in, ws_recruits, ws_nofos):
    columns_in = get_column_mapping(ws_in)
    columns_recruits = get_column_mapping(ws_recruits)
    columns_nofos = get_column_mapping(ws_nofos)

    row_in = 2
    row_recruits = 2
    row_nofos = 2
    row_in_max = ws_in.max_row
    while row_in <= row_in_max:
        if ws_in.cell(row=row_in, column=columns_in["Title Description"]).value.strip() == "65 - 65 (BY CONVERSION)":
            row_in += 1
            continue
        
        temp_columns = columns_recruits
        temp_ws = ws_recruits
        temp_row = row_recruits
        nofos = ws_in.cell(row=row_in, column=columns_in["BP per Product"]).value == 1
        if nofos:
            temp_columns = columns_nofos
            temp_ws = ws_nofos
            temp_row = row_nofos
            
        for column, index in temp_columns.items():
            input_value = ws_in.cell(row=row_in, column=columns_in[column]).value
            temp_ws.cell(row=temp_row, column=index).value = input_value

        row_in += 1

        if nofos:
            row_nofos += 1
        else:
            row_recruits += 1


def Make_TitlesReports_Mapping(TitlesReportMatureMarketsWB):
    mapping = {}
    # 98123: {
    #   "Member group": "Guest Customer",
    #   "Director First Name": "JACKIE",
    #   "Director Last Name": "VENABLES",
    # }
    ws = TitlesReportMatureMarketsWB["C1-M2 titles all"]
    columns = get_column_mapping(ws, row_with_columns=11)
    
    for row in range(12, ws.max_row):
        consultant_number = ws.cell(row=row, column=columns["Consultant (Downline)"]).value
        if consultant_number not in mapping:
            member_group = ws.cell(row=row, column=columns["Legacy"]).value
            mapping[consultant_number] = {}
            mapping[consultant_number]["Member group"] = member_group

            director_first_name = ws.cell(row=row, column=columns["Director First Name"]).value
            mapping[consultant_number]["Director First Name"] = director_first_name
            director_last_name = ws.cell(row=row, column=columns["Director Last Name"]).value
            mapping[consultant_number]["Director Last Name"] = director_last_name


    ws = TitlesReportMatureMarketsWB["VIP with BP >0"]
    columns = get_column_mapping(ws, row_with_columns=12)

    for row in range(13, ws.max_row):
        consultant_number = ws.cell(row=row, column=columns["End Customer"]).value
        if consultant_number not in mapping:
            member_group = ws.cell(row=row, column=columns["End Customer Type"]).value
            mapping[consultant_number] = {}
            mapping[consultant_number]["Member group"] = member_group

            director_first_name = ws.cell(row=row, column=columns["Director First Name"]).value
            mapping[consultant_number]["Director First Name"] = director_first_name
            director_last_name = ws.cell(row=row, column=columns["Director Last Name"]).value
            mapping[consultant_number]["Director Last Name"] = director_last_name
    
    return mapping


def copy_skincare_case(ws_out, ws_in, mapping):
    output_columns = get_column_mapping(ws_out)
    input_columns = get_column_mapping(ws_in)

    row_in = 2
    row_out = 2
    row_in_max = ws_in.max_row
    while row_in <= row_in_max:
        if ws_in.cell(row=row_in, column=input_columns["Title Description"]).value.strip() == "65 - 65 (BY CONVERSION)":
            row_in += 1
            continue
            
        for column, index in output_columns.items():
            if column not in input_columns:
                continue

            input_value = ws_in.cell(row=row_in, column=input_columns[column]).value
            ws_out.cell(row=row_out, column=index).value = input_value

            if column == "Consultant number":
                if input_value in mapping:
                    member_group = mapping[input_value]["Member group"]
                    ws_out.cell(row=row_out, column=output_columns["Member group"]).value = member_group
                
        row_in += 1 
        row_out += 1


def copy_catalogue_titles_case(ws_out1, ws_out2, wb_in):
    ws_in = wb_in["C1-M2 titles all"]

    output_columns = get_column_mapping(ws_out1)
    input_columns = get_column_mapping(ws_in, row_with_columns=11)

    row_out = 2
    for row in range(12, ws_in.max_row+1):
            
        for column, index in output_columns.items():
            input_value = ws_in.cell(row=row, column=input_columns[column]).value
            ws_out1.cell(row=row_out, column=index).value = input_value
            ws_out2.cell(row=row_out, column=index).value = input_value
        
        row_out += 1
        

def copy_YTD_case(ws_out, ws_in, mapping): 
    output_columns = get_column_mapping(ws_out)
    input_columns = get_column_mapping(ws_in)

    for row in range(2, ws_in.max_row+1):    
        for column, index in output_columns.items():
            
            if column not in input_columns:
                continue

            input_value = ws_in.cell(row=row, column=input_columns[column]).value
            ws_out.cell(row=row, column=index).value = input_value

            if column == "CONSULTANT":
                if input_value in mapping:
                    director_first_name = mapping[input_value]["Director First Name"]
                    ws_out.cell(row=row, column=output_columns["DIRECTOR_FIRST_NAME"]).value = director_first_name
                    director_last_name = mapping[input_value]["Director Last Name"]
                    ws_out.cell(row=row, column=output_columns["DIRECTOR_LAST_NAME"]).value = director_last_name


def copy_VIP_Recruits_case(ws_out, ws_in, date_from, date_to):

    output_columns = get_column_mapping(ws_out)
    input_columns = get_column_mapping(ws_in, row_with_columns=12)

    row_out = 2
    row_in = 13
    row_in_max = ws_in.max_row
    while row_in <= row_in_max:
        
        end_customer_type = ws_in.cell(row=row_in, column=input_columns["End Customer Type"]).value
        if end_customer_type != "VIP Customer":
            row_in += 1
            continue
        
        order_date = ws_in.cell(row=row_in, column=input_columns["End customers first order date"]).value
        if not (date_from <= order_date and order_date <= date_to):
            row_in += 1
            continue

        for column, index in output_columns.items():
            input_value = ws_in.cell(row=row_in, column=input_columns[column]).value
            ws_out.cell(row=row_out, column=index).value = input_value
        
        row_in += 1
        row_out += 1


# TODO: filenames argument is useless
def copy_all(input_directory: str, filenames: List[str], path_to_save: str, date_from: datetime, date_to: datetime, set_progress_text_fn: Callable[[str], None]) -> None:
    
    set_progress_text_fn("Loading Directors workbook...")
    output_workbook = xl.load_workbook(input_directory + "/Directors Basic working file.xlsm", keep_vba=True)

    ######################################
    ### Copy simple cases

    simple_cases = [
        "Signups",
        "Starter Kits",
        "Welcome Programme",
        "I3"
    ]

    for case in simple_cases:
        set_progress_text_fn(case)
        wb_in = xl.load_workbook(input_directory + f"/{case}.xlsx")
        copy_simple_case(output_workbook[case], wb_in.active)
    

    ######################################
    ### Copy "Recruits (FO)" and "NOFOs who paid joining fee" case

    set_progress_text_fn("Recruits (FO) and NOFOs")
    wb_in = xl.load_workbook(input_directory + "/Recruits (FO).xlsx")
    copy_recruits(wb_in.active, output_workbook["Recruits (FO)"], output_workbook["NOFOs who paid joining fee"])
    
    ######################################
    ### Copy "Skincare Sets" case
    
    set_progress_text_fn("Skincare Sets")
    # Make "Consultant number - Member group" mapping
    titles_reports_wb = xl.load_workbook(input_directory + "/Titles Report Mature Markets.xlsx")
    mapping = Make_TitlesReports_Mapping(titles_reports_wb)

    wb_in = xl.load_workbook(input_directory + "/Skincare Sets.xlsx")
    copy_skincare_case(output_workbook["Skincare Sets"], wb_in.active, mapping)

    
    ######################################
    ### Copy "Catalogue BP Sales" oraz "Titles" cases

    set_progress_text_fn("Catalogue BP Sales and Titles")
    copy_catalogue_titles_case(
        output_workbook["Catalogue BP Sales"],
        output_workbook["Titles"],
        titles_reports_wb
    )


    ######################################
    ### Copy "YTD"

    set_progress_text_fn("YTD")
    ws_in = xl.load_workbook(input_directory + "/YTD.xlsx")
    copy_YTD_case(output_workbook["YTD"], ws_in.active, mapping)


    ######################################
    ### Copy "VIP Recruits"

    set_progress_text_fn("VIP Recruits")

    copy_VIP_Recruits_case(output_workbook["VIP Recruits"], titles_reports_wb["VIP with BP >0"], date_from, date_to)



    set_progress_text_fn("Saving now...")
    output_workbook.save(path_to_save)
    set_progress_text_fn("Done!")

